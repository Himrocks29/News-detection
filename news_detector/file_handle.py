import openpyxl

def handle(head, res):
	wb = openpyxl.load_workbook('D:/News-detection-master/News-detection-master/news_detector/database.xlsx')
	ws = wb['Sheet1']
	i = ws.max_row
	ws.cell(row = i+1, column = 1).value = (i-1)
	ws.cell(row = i+1, column = 2).value = head
	ws.cell(row = i+1, column = 3).value = res

	wb.save('D:/News-detection-master/News-detection-master/news_detector/database.xlsx')
	wb.close()

#handle("China shake hands with pakristan", "9")

